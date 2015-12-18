import openpyxl
import os
import requests
from bs4 import BeautifulSoup

# 1.0 Read urls from the excel file
# 1.1 Change the current path of the file
path = "C:\\Users\\maitreypatel\\Desktop\\TaxonomyUpdates\\LATTE-931-Pets-Taxonomy"
workBook = "Pets_Taxonomy_Update_V1.3.xlsx"
sheetNameRename = "Rename Validation"
sheetNameInsert = "Insert Validation"
sheetNameDelete = "Delete Validation"

def changeDirectory():
    changedir = os.chdir(path)
excelNames = []
getCrumb = []
# 1.2 Go to the excel file where the taxonomies to be validated
def getRenames():
    getRenames = openpyxl.load_workbook(workBook).get_sheet_by_name(sheetNameRename).columns[8]
    for renames in getRenames:
        names = renames.value
        req = requests.get(names)
        soup = BeautifulSoup(req.content, 'html.parser')
        crumbList=soup.find("div", {"id": "brd-crumbs"}).find_all("span")[0].text
        getCrumb.append(crumbList)
    #print(getCrumb)

def getFunnyNames():
    getExcelNames = openpyxl.load_workbook(workBook).get_sheet_by_name(sheetNameRename).columns[0]
    for taxRenames in getExcelNames:
        excelNames.append(taxRenames.value)
    #print(excelNames)
    #return excelNames

# Compare the parsed taxonomy values with the excel taxonomy values
notMatchedRenames = 0
def compareNames():
    for lastCrumbs in getCrumb:
        if lastCrumbs not in excelNames:
            print(lastCrumbs)
            notMatchedRenames=+1
        if notMatchedRenames==0:
            print("All renames are matching")

invalidInserts = 0
def verifyInserts():
    getInsertNames = openpyxl.load_workbook(workBook).get_sheet_by_name(sheetNameInsert).columns[6]
    for inserts in getInsertNames:
        insertNames = inserts.value
        req_status = requests.get(insertNames).status_code
        if req_status != 200:
            print(insertNames)
            invalidInserts=+1
    print(invalidInserts)
    if invalidInserts==0:
        print("All insert urls are correct")

invalidDeleteCounts = 0
def verifyDeletes():
    getDeletedNames = openpyxl.load_workbook(workBook).get_sheet_by_name(sheetNameDelete).columns[6]






changeDirectory()
#getRenames()
#getFunnyNames()
#compareNames()
verifyInserts()







# Pets_Taxonomy_Update_V1.3.xls

# Save the taxonomies name from the excel file

# Parse the names of the taxonomies from the page sources


